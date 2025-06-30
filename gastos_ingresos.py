import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import fitz  # PyMuPDF
import os
import csv
from datetime import datetime
import re

def obtener_archivo_mes(mes):
    return f"movimientos_{mes}.csv"

def inicializar_archivo_mes(mes):
    archivo = obtener_archivo_mes(mes)
    if not os.path.exists(archivo):
        with open(archivo, "w", newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Fecha", "Tipo", "Descripción", "Monto", "Categoría"])

def guardar_movimiento(tipo, descripcion, monto, categoria="General", fecha=None, mes=None):
    try:
        if mes is None:
            mes = datetime.now().strftime("%Y-%m")
        archivo = obtener_archivo_mes(mes)
        with open(archivo, "a", newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            fecha_str = datetime.now().strftime("%Y-%m-%d %H:%M") if fecha is None else fecha.strftime("%Y-%m-%d %H:%M")
            writer.writerow([fecha_str, tipo, descripcion, monto, categoria])
        actualizar_saldo()
        actualizar_tabla()
        return True
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el movimiento.\nError: {str(e)}")
        return False

def leer_csv_mes(mes):
    archivo = obtener_archivo_mes(mes)
    if not os.path.exists(archivo):
        return pd.DataFrame()
    try:
        df = pd.read_csv(archivo, encoding='utf-8')
        return df
    except:
        return pd.DataFrame()

def calcular_total():
    try:
        df = leer_csv_mes(mes_var.get())
        if df.empty:
            return 0.0
        df["Monto"] = pd.to_numeric(df["Monto"], errors='coerce')
        ingresos = df[df["Tipo"] == "Ingreso"]["Monto"].sum()
        gastos = df[df["Tipo"] == "Retiro"]["Monto"].sum()
        return float(ingresos - gastos)
    except:
        return 0.0

def actualizar_saldo():
    total = calcular_total()
    saldo_label.config(text=f"Saldo {mes_var.get()}: {total:.2f} €")
    saldo_label.config(fg="red" if total < 0 else "green")

def actualizar_tabla(filtro=None):
    for item in tree.get_children():
        tree.delete(item)
    try:
        df = leer_csv_mes(mes_var.get())
        if df.empty:
            return
        df["Monto"] = pd.to_numeric(df["Monto"], errors='coerce')
        df["Fecha"] = pd.to_datetime(df["Fecha"])
        df = df.sort_values("Fecha", ascending=False)
        if filtro:
            mask = df["Descripción"].str.contains(filtro, case=False, na=False) |                    df["Tipo"].str.contains(filtro, case=False, na=False) |                    df["Fecha"].astype(str).str.contains(filtro, case=False, na=False) |                    df["Categoría"].str.contains(filtro, case=False, na=False)
            df = df[mask]
        for _, row in df.iterrows():
            tree.insert("", "end", values=(
                row["Fecha"].strftime("%Y-%m-%d %H:%M"),
                row["Tipo"],
                row["Descripción"],
                f"{row['Monto']:.2f} €",
                row["Categoría"]
            ))
    except Exception as e:
        print(f"Error actualizando tabla: {str(e)}")

def buscar_movimientos():
    texto_busqueda = busqueda_entry.get().strip()
    actualizar_tabla(texto_busqueda if texto_busqueda else None)

def registrar_manual(tipo):
    def guardar():
        desc = descripcion_entry.get()
        monto = monto_entry.get()
        cat = categoria_entry.get()
        try:
            monto_float = float(monto)
        except:
            messagebox.showerror("Error", "Monto inválido")
            return
        if guardar_movimiento(tipo, desc, monto_float, cat, mes=mes_var.get()):
            ventana_registro.destroy()

    ventana_registro = tk.Toplevel(ventana)
    ventana_registro.title(f"Registrar {tipo}")
    ventana_registro.geometry("400x250")
    ventana_registro.configure(bg="#f0f0f0")

    tk.Label(ventana_registro, text="Descripción:", bg="#f0f0f0").pack(pady=5)
    descripcion_entry = tk.Entry(ventana_registro, width=40)
    descripcion_entry.pack(pady=5)

    tk.Label(ventana_registro, text="Monto:", bg="#f0f0f0").pack(pady=5)
    monto_entry = tk.Entry(ventana_registro, width=20)
    monto_entry.pack(pady=5)

    tk.Label(ventana_registro, text="Categoría:", bg="#f0f0f0").pack(pady=5)
    categoria_entry = tk.Entry(ventana_registro, width=30)
    categoria_entry.insert(0, "General")
    categoria_entry.pack(pady=5)

    tk.Button(ventana_registro, text="Guardar", command=guardar, bg="#2ecc71", fg="white").pack(pady=15)

def importar_pdf():
    archivo = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if not archivo:
        return
    try:
        doc = fitz.open(archivo)
        texto = ""
        for page in doc:
            texto += page.get_text()
        doc.close()

        match = re.search(r"L[ií]quido\s+a\s+percibir\s*[:\-]?\s*(\d+[.,]?\d*)", texto, re.IGNORECASE)
        if match:
            monto = float(match.group(1).replace(",", "."))
        else:
            posibles = re.findall(r"\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})", texto)
            montos = [float(p.replace(".", "").replace(",", ".")) for p in posibles if 100 <= float(p.replace(".", "").replace(",", ".")) <= 10000]
            monto = max(montos) if montos else 0

        if monto > 0:
            if guardar_movimiento("Ingreso", f"Nómina PDF - {os.path.basename(archivo)}", monto, "Nómina", mes=mes_var.get()):
                messagebox.showinfo("Éxito", f"Importado correctamente: {monto:.2f} €")
        else:
            messagebox.showerror("Error", "No se pudo detectar el líquido a percibir.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el movimiento.\nError: {str(e)}")


def exportar_csv():
    archivo = filedialog.asksaveasfilename(defaultextension=".csv",
        filetypes=[("CSV files", "*.csv")],
        title="Exportar CSV")
    if archivo:
        df = leer_csv_mes(mes_var.get())
        if df.empty:
            messagebox.showinfo("Aviso", "No hay datos para exportar.")
            return
        df.to_csv(archivo, index=False, encoding='utf-8')
        messagebox.showinfo("Exportado", f"Datos exportados a {archivo}")
def eliminar_movimiento():
    seleccion = tree.selection()
    if not seleccion:
        messagebox.showwarning("Advertencia", "Selecciona un movimiento para eliminar.")
        return

    confirm = messagebox.askyesno("Confirmación", "¿Eliminar el movimiento seleccionado?")
    if not confirm:
        return

    item = tree.item(seleccion[0])
    valores = item['values']
    try:
        fecha, tipo, descripcion, monto_str, categoria = valores
        monto = float(monto_str.replace(" €", ""))
        df = leer_csv_mes(mes_var.get())
        df_filtrado = df[~(
            (df["Fecha"] == fecha) &
            (df["Tipo"] == tipo) &
            (df["Descripción"] == descripcion) &
            (abs(df["Monto"] - monto) < 0.01) &
            (df["Categoría"] == categoria)
        )]
        archivo = obtener_archivo_mes(mes_var.get())
        df_filtrado.to_csv(archivo, index=False, encoding='utf-8')
        messagebox.showinfo("Éxito", "Movimiento eliminado.")
        actualizar_tabla()
        actualizar_saldo()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo eliminar.\n{str(e)}")

def exportar_excel():
    archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Exportar Excel")
    if not archivo:
        return

    df = leer_csv_mes(mes_var.get())
    if df.empty:
        messagebox.showwarning("Advertencia", "No hay datos para exportar.")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)
            for c_idx, cell in enumerate(ws[r_idx], start=1):
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        monto_idx = df.columns.get_loc("Monto") + 1
        for row in ws.iter_rows(min_row=2, min_col=monto_idx, max_col=monto_idx):
            for cell in row:
                cell.number_format = '€#,##0.00'

        if "Fecha" in df.columns:
            fecha_idx = df.columns.get_loc("Fecha") + 1
            for row in ws.iter_rows(min_row=2, min_col=fecha_idx, max_col=fecha_idx):
                for cell in row:
                    cell.number_format = "YYYY-MM-DD HH:MM"

        ws.auto_filter.ref = ws.dimensions

        wb.save(archivo)
        messagebox.showinfo("Éxito", f"Archivo Excel exportado:\n{archivo}")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a Excel.\nError: {str(e)}")
# GUI
ventana = tk.Tk()
ventana.title("Control de Gastos Mensuales")
ventana.geometry("950x600")
ventana.configure(bg="#f0f0f0")

# Mes seleccionado
mes_actual = datetime.now().strftime("%Y-%m")
ANIO_ACTUAL = datetime.now().year
meses_disponibles = [f"{y}-{m:02d}" for y in range(2023, ANIO_ACTUAL + 2) for m in range(1, 13)]
mes_var = tk.StringVar(value=mes_actual)

tk.Label(ventana, text="Selecciona el mes:", bg="#f0f0f0").pack()
mes_menu = ttk.Combobox(ventana, textvariable=mes_var, values=meses_disponibles, width=10, state="readonly")
mes_menu.pack()
mes_menu.bind("<<ComboboxSelected>>", lambda e: (inicializar_archivo_mes(mes_var.get()), actualizar_saldo(), actualizar_tabla()))

saldo_label = tk.Label(ventana, text="Saldo: 0.00 €", font=("Arial", 14, "bold"), bg="#f0f0f0", fg="green")
saldo_label.pack(pady=10)

frame_botones = tk.Frame(ventana, bg="#f0f0f0")
frame_botones.pack(pady=5)

tk.Button(frame_botones, text="Registrar Ingreso", command=lambda: registrar_manual("Ingreso"),
          bg="#2ecc71", fg="white", width=20).pack(side="left", padx=5)
tk.Button(frame_botones, text="Registrar Gasto", command=lambda: registrar_manual("Retiro"),
          bg="#e74c3c", fg="white", width=20).pack(side="left", padx=5)
tk.Button(frame_botones, text="Importar Nómina PDF", command=importar_pdf,
          bg="#2980b9", fg="white", width=20).pack(side="left", padx=5)
tk.Button(frame_botones, text="Exportar CSV", command=exportar_csv,
          bg="#8e44ad", fg="white", width=20).pack(side="left", padx=5)
tk.Button(frame_botones, text="📈 Exportar Excel", command=exportar_excel,
          bg="#1abc9c", fg="white", width=20).pack(side="left", padx=5)
tk.Button(frame_botones, text="Eliminar Movimiento", command=eliminar_movimiento,
          bg="#c0392b", fg="white", width=20).pack(side="left", padx=5)
frame_busqueda = tk.Frame(ventana, bg="#f0f0f0")
frame_busqueda.pack(pady=5)
busqueda_entry = tk.Entry(frame_busqueda, width=50)
busqueda_entry.pack(side="left", padx=5)
tk.Button(frame_busqueda, text="Buscar", command=buscar_movimientos,
          bg="#16a085", fg="white").pack(side="left")

frame_tabla = tk.Frame(ventana)
frame_tabla.pack(fill="both", expand=True, padx=10, pady=10)

columns = ("Fecha", "Tipo", "Descripción", "Monto", "Categoría")
tree = ttk.Treeview(frame_tabla, columns=columns, show="headings", height=15)
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=140 if col == "Fecha" else 120)
tree.pack(side="left", fill="both", expand=True)

scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side="right", fill="y")

inicializar_archivo_mes(mes_var.get())
actualizar_saldo()
actualizar_tabla()
ventana.mainloop()
